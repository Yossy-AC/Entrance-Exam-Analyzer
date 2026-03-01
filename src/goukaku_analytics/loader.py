from pathlib import Path
import logging
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

logger = logging.getLogger("goukaku_analytics")

# ファイルパス+更新時刻ベースのキャッシュ（同一ファイルの繰り返し読み込み防止）
_cache: dict[str, tuple[float, pd.DataFrame]] = {}

# 分類コードのマッピング
CATEGORY_LABELS = {
    1: "国立", 2: "公立", 3: "大学校", 4: "私立", 5: "看護専門",
    6: "その他専門", 7: "専門職大学", 8: "短大", 9: "就職", 10: "浪人", 11: "不明",
}

# 合否コードのマッピング
PASS_LABELS = {0: "出願", 1: "合格", 2: "不合格"}

# 列インデックス → 標準列名のマッピング（固定位置）
COL_MAP = {
    0: "No",
    1: "生徒コード",
    2: "教室",
    3: "学年",
    4: "性別",
    5: "学校",
    6: "学科",
    7: "JA併用",
    8: "通塾履歴",
    9: "VIP",
    10: "部活",
    11: "承諾書",
    12: "撮影",
    13: "合格アンケ",
    14: "更新日",
    15: "合格",
    16: "合格状態",
    17: "進学先決定",
    18: "進学先状態",
    19: "分類",
    20: "分類補足",
    21: "大学名",
    22: "学部名",
    23: "学科名",
    24: "志望順位",
    25: "方式",
    26: "方式補足",
    27: "発表日",
    28: "備考",
    # 第2志望（実Excelに「分類2補足」列は存在しない）
    29: "分類2",
    30: "大学名2",
    31: "学部名2",
    32: "学科名2",
    33: "備考2",
    # 第3志望（実Excelに「分類3補足」列は存在しない）
    34: "分類3",
    35: "大学名3",
    36: "学部名3",
    37: "学科名3",
    38: "備考3",
    # 共通テスト得点
    39: "合計得点",
    40: "得点率",
    41: "満点",
    42: "文理区分",
    43: "英R",
    44: "英L",
    45: "数学1",
    46: "数学2",
    47: "国語",
    48: "物理",
    49: "化学",
    50: "生物",
    51: "地学",
    52: "物理基礎",
    53: "化学基礎",
    54: "生物基礎",
    55: "地学基礎",
    56: "世界史",
    57: "日本史",
    58: "地理",
    59: "倫理",
    60: "政経",
    61: "地歴公共",
    62: "情報",
}

SUBJECT_COLUMNS = [
    "英R", "英L", "数学1", "数学2", "国語",
    "物理", "化学", "生物", "地学",
    "物理基礎", "化学基礎", "生物基礎", "地学基礎",
    "世界史", "日本史", "地理", "倫理", "政経", "地歴公共", "情報",
]


def _find_header_row_idx(ws) -> int:
    """'No' が含まれるヘッダー行の行番号(1始まり)を返す"""
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        if any(str(v).strip() == "No" for v in row if v is not None):
            return i + 1  # 1始まりに変換
    return 5


def load_data(path: Path) -> pd.DataFrame:
    """Excelからメインデータシートを読み込みDataFrameとして返す（mtime ベースキャッシュ付き）"""
    key = str(path.resolve())
    mtime = path.stat().st_mtime
    if key in _cache and _cache[key][0] == mtime:
        return _cache[key][1].copy()

    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    header_row = _find_header_row_idx(ws)
    data_start = header_row + 1

    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        rows.append(list(row))

    max_cols = max(len(r) for r in rows) if rows else 70
    col_names = [COL_MAP.get(i, f"col_{i}") for i in range(max_cols)]

    # 短い行をパディング
    padded = [r + [None] * (max_cols - len(r)) for r in rows]
    df = pd.DataFrame(padded, columns=col_names[:max_cols])

    # 空行（No が None）を除外
    df = df.dropna(subset=["No"]).copy()
    df["No"] = pd.to_numeric(df["No"], errors="coerce")
    df = df.dropna(subset=["No"]).copy()
    df["No"] = df["No"].astype(int)

    # 合否フラグ
    df["合格"] = pd.to_numeric(df["合格"], errors="coerce").fillna(0).astype(int)
    df["合格ラベル"] = df["合格"].map(PASS_LABELS).fillna("不明")

    # 分類ラベル
    df["分類"] = pd.to_numeric(df["分類"], errors="coerce")
    df["分類ラベル"] = df["分類"].map(CATEGORY_LABELS).fillna("不明")

    # 国公私区分
    df["国公私"] = df["分類"].apply(
        lambda x: "国公立" if x in [1, 2, 3] else ("私立等" if x in [4, 5, 6, 7, 8] else "その他")
    )

    # 文理区分
    df["文理区分"] = pd.to_numeric(df["文理区分"], errors="coerce")
    df["文理"] = df["文理区分"].map({1: "文系", 2: "理系"}).fillna("不明")

    # 志望順位・進学先決定を数値化
    df["志望順位"] = pd.to_numeric(df["志望順位"], errors="coerce")
    df["進学先決定"] = pd.to_numeric(df["進学先決定"], errors="coerce")

    # 共通テスト得点
    df["合計得点"] = pd.to_numeric(df["合計得点"], errors="coerce")
    df["得点率"] = pd.to_numeric(df["得点率"], errors="coerce")

    # 科目得点を数値変換
    for col in SUBJECT_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    _cache[key] = (mtime, df)
    return df.copy()


_EXCEL_EPOCH = datetime(1899, 12, 30)


def get_update_date(df: pd.DataFrame) -> str:
    """データの更新日列（COL_MAP 14列目）の最新値を返す"""
    try:
        col = df["更新日"].dropna()
        if len(col) == 0:
            return "不明"
        latest = col.max()
        if isinstance(latest, datetime):
            return latest.strftime("%Y-%m-%d")
        if isinstance(latest, (int, float)):
            return (_EXCEL_EPOCH + timedelta(days=int(latest))).strftime("%Y-%m-%d")
        return str(latest)
    except Exception:
        return "不明"


def load_all_years(year_paths: dict[int, Path]) -> dict[int, pd.DataFrame]:
    """複数年度のデータを読み込む"""
    result = {}
    for year, path in year_paths.items():
        try:
            result[year] = load_data(path)
        except Exception as e:
            logger.warning("%s年度データの読み込みに失敗: %s", year, e)
    return result
